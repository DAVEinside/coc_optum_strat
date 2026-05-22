"""Parser for `data_repo/structs_data.xlsx` — wiki-sourced building upgrade data.

Layout: each building has a name row, then a header row, then per-level data rows,
then a blank row separating from the next building.

Headers vary across buildings (different column orders, sub-header rows for
multi-mode towers), so we identify cost/time/TH columns by header *text* rather
than fixed position.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd


# --- Build time parsing ---
_TIME_UNITS = {"d": 86400, "h": 3600, "m": 60, "s": 1}
_TIME_RE = re.compile(r"(\d+)\s*([dhms])")


def parse_time_string(s) -> Optional[int]:
    """'1d 12h' -> 129600 ; '5s' -> 5 ; 'N/A' -> None."""
    if s is None:
        return None
    s = str(s).strip()
    if not s or s.upper() == "N/A":
        return None
    total = 0
    for amt, unit in _TIME_RE.findall(s.lower()):
        total += int(amt) * _TIME_UNITS[unit]
    return total if total > 0 else None


def parse_cost(s) -> Optional[int]:
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return int(s)
    s = str(s).strip()
    if not s or s.upper() == "N/A":
        return None
    try:
        return int(float(s.replace(",", "")))
    except ValueError:
        return None


# --- Header column identification ---
def _norm(text) -> str:
    return str(text).strip().replace("\xa0", " ").lower() if text is not None else ""


def _find_col(headers: List, candidates: List[str]) -> Optional[int]:
    """Return index of the first header that exactly matches any candidate (case-insensitive)."""
    norm_headers = [_norm(h) for h in headers]
    cand = [c.lower() for c in candidates]
    for i, h in enumerate(norm_headers):
        if h in cand:
            return i
    return None


COST_HEADERS = ["cost", "build cost"]
TIME_HEADERS = ["build time"]
TH_HEADERS = ["town hall level required", "th level required"]
LEVEL_HEADERS = ["level", "th level"]
CAPACITY_HEADERS = ["storage capacity", "capacity"]


# --- Resource mapping per building (Gold/Elixir/Dark_Elixir) ---
# Source: standard CoC knowledge. "Cost" column in xlsx doesn't say which resource.
RESOURCE_BY_BUILDING: Dict[str, str] = {
    # Defenses: Gold
    "cannon": "gold", "archer tower": "gold", "mortar": "gold", "air defense": "gold",
    "wizard tower": "gold", "air sweeper": "gold", "hidden tesla": "gold",
    "bomb tower": "gold", "x-bow": "gold", "inferno tower": "gold",
    "eagle artillery": "gold", "scattershot": "gold", "builder hut": "gold",
    "spell tower": "gold", "monolith": "gold", "multi archer tower": "gold",
    "ricochet cannon": "gold", "multi-gear tower": "gold",
    "firespitter": "gold", "revenge tower": "gold",
    "super wizard tower": "gold", "roaster": "gold", "airbomb": "gold",
    "lava launcher": "gold",
    # Traps: Gold (most), DE for some at higher tiers (we'll lump as gold for simplicity)
    "walls": "gold", "bombs": "gold", "spring traps": "gold", "giant bomb": "gold",
    "air bomb": "gold", "seeking air mine": "gold", "skeleton trap": "gold",
    "tornado trap": "gold", "giga bomb": "gold",
    # Town Hall: Gold
    "town hall": "gold",
    # Resource buildings: Mixed (Gold buildings cost Elixir to upgrade & vice versa)
    "gold mine": "elixir", "elixar collector": "gold", "gold storage": "elixir",
    "elixar storage": "gold", "dark elixar drill": "gold", "dark elixar storage": "gold",
    # Army buildings: Elixir
    "clan castle": "gold", "army camp": "elixir", "barracks": "elixir",
    "dark barracks": "elixir", "laboratory": "elixir", "spell factory": "elixir",
    "hero hall": "elixir", "dark spell factory": "elixir",
    "blachsmith": "elixir", "workshop": "elixir", "pet house": "elixir",
}


# --- Instance counts per (building, TH level) ---
# We track building counts at each TH. Format: building_name (lowercase) -> {th_level: count}.
# Numbers reflect community-confirmed defense layouts (Fandom wiki "Buildings Available" pages).
INSTANCE_COUNT: Dict[str, Dict[int, int]] = {
    "town hall": {th: 1 for th in range(1, 19)},
    "cannon": {1: 2, 2: 2, 3: 2, 4: 2, 5: 3, 6: 3, 7: 5, 8: 5, 9: 6, 10: 6, 11: 7, 12: 7, 13: 7, 14: 7, 15: 7, 16: 7, 17: 7, 18: 7},
    "archer tower": {2: 1, 3: 1, 4: 2, 5: 3, 6: 4, 7: 5, 8: 6, 9: 7, 10: 8, 11: 8, 12: 8, 13: 8, 14: 8, 15: 8, 16: 8, 17: 8, 18: 8},
    "mortar": {3: 1, 4: 1, 5: 1, 6: 2, 7: 3, 8: 4, 9: 4, 10: 4, 11: 4, 12: 4, 13: 4, 14: 4, 15: 4, 16: 4, 17: 4, 18: 4},
    "air defense": {4: 1, 5: 2, 6: 2, 7: 3, 8: 3, 9: 4, 10: 4, 11: 4, 12: 4, 13: 4, 14: 4, 15: 4, 16: 4, 17: 4, 18: 4},
    "wizard tower": {5: 1, 6: 1, 7: 2, 8: 3, 9: 4, 10: 4, 11: 5, 12: 5, 13: 5, 14: 5, 15: 5, 16: 5, 17: 5, 18: 5},
    "air sweeper": {6: 1, 7: 1, 8: 1, 9: 1, 10: 2, 11: 2, 12: 2, 13: 2, 14: 2, 15: 2, 16: 2, 17: 2, 18: 2},
    "hidden tesla": {7: 2, 8: 3, 9: 4, 10: 4, 11: 5, 12: 5, 13: 6, 14: 6, 15: 7, 16: 7, 17: 7, 18: 7},
    "bomb tower": {8: 1, 9: 2, 10: 2, 11: 2, 12: 2, 13: 2, 14: 2, 15: 2, 16: 2, 17: 2, 18: 2},
    "x-bow": {9: 2, 10: 3, 11: 3, 12: 4, 13: 4, 14: 4, 15: 4, 16: 4, 17: 4, 18: 4},
    "inferno tower": {10: 2, 11: 2, 12: 3, 13: 3, 14: 3, 15: 3, 16: 3, 17: 3, 18: 3},
    "eagle artillery": {11: 1, 12: 1, 13: 1, 14: 1, 15: 1, 16: 1, 17: 1, 18: 1},
    "scattershot": {13: 2, 14: 2, 15: 4, 16: 4, 17: 4, 18: 4},
    "builder hut": {1: 2, 2: 3, 3: 3, 4: 4, 5: 5, 6: 5, 7: 5, 8: 5, 9: 5, 10: 5, 11: 5, 12: 5, 13: 5, 14: 5, 15: 5, 16: 5, 17: 5, 18: 5},
    "spell tower": {15: 2, 16: 2, 17: 2, 18: 2},
    "monolith": {15: 1, 16: 1, 17: 1, 18: 1},
    "multi archer tower": {16: 1, 17: 1, 18: 1},
    "ricochet cannon": {16: 2, 17: 2, 18: 2},
    "multi-gear tower": {16: 1, 17: 1, 18: 1},
    "firespitter": {17: 1, 18: 1},
    "revenge tower": {17: 1, 18: 1},
    "super wizard tower": {18: 1},
    "roaster": {18: 1},
    "airbomb": {18: 1},
    "lava launcher": {18: 1},
    # Traps (only the count active at each TH, not really a "building" in scheduling sense)
    "walls": {3: 25, 4: 50, 5: 75, 6: 75, 7: 75, 8: 100, 9: 175, 10: 225, 11: 250, 12: 275, 13: 275, 14: 300, 15: 325, 16: 325, 17: 325, 18: 325},
    "bombs": {3: 2, 4: 4, 5: 6, 6: 6, 7: 6, 8: 6, 9: 6, 10: 6, 11: 6, 12: 6, 13: 8, 14: 8, 15: 10, 16: 10, 17: 12, 18: 12},
    "spring traps": {4: 2, 5: 2, 6: 4, 7: 4, 8: 4, 9: 6, 10: 6, 11: 6, 12: 6, 13: 6, 14: 6, 15: 6, 16: 6, 17: 6, 18: 6},
    "giant bomb": {5: 1, 6: 2, 7: 2, 8: 3, 9: 3, 10: 4, 11: 5, 12: 5, 13: 6, 14: 7, 15: 7, 16: 7, 17: 7, 18: 7},
    "air bomb": {6: 1, 7: 2, 8: 3, 9: 4, 10: 5, 11: 5, 12: 5, 13: 6, 14: 6, 15: 6, 16: 6, 17: 6, 18: 6},
    "seeking air mine": {8: 1, 9: 1, 10: 2, 11: 3, 12: 4, 13: 5, 14: 5, 15: 6, 16: 6, 17: 6, 18: 6},
    "skeleton trap": {10: 2, 11: 3, 12: 4, 13: 5, 14: 6, 15: 6, 16: 6, 17: 6, 18: 6},
    "tornado trap": {12: 1, 13: 2, 14: 2, 15: 2, 16: 2, 17: 2, 18: 2},
    "giga bomb": {16: 1, 17: 1, 18: 1},
    # Resource buildings
    "gold mine": {1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 5, 7: 6, 8: 6, 9: 7, 10: 7, 11: 7, 12: 7, 13: 7, 14: 7, 15: 7, 16: 7, 17: 7, 18: 7},
    "elixar collector": {1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 5, 7: 6, 8: 6, 9: 7, 10: 7, 11: 7, 12: 7, 13: 7, 14: 7, 15: 7, 16: 7, 17: 7, 18: 7},
    "gold storage": {1: 1, 2: 2, 3: 2, 4: 2, 5: 2, 6: 2, 7: 2, 8: 3, 9: 4, 10: 4, 11: 4, 12: 4, 13: 4, 14: 4, 15: 4, 16: 4, 17: 4, 18: 4},
    "elixar storage": {1: 1, 2: 2, 3: 2, 4: 2, 5: 2, 6: 2, 7: 2, 8: 3, 9: 4, 10: 4, 11: 4, 12: 4, 13: 4, 14: 4, 15: 4, 16: 4, 17: 4, 18: 4},
    "dark elixar drill": {8: 1, 9: 2, 10: 3, 11: 3, 12: 3, 13: 3, 14: 3, 15: 3, 16: 3, 17: 3, 18: 3},
    "dark elixar storage": {7: 1, 8: 1, 9: 1, 10: 1, 11: 1, 12: 1, 13: 1, 14: 1, 15: 1, 16: 1, 17: 1, 18: 1},
    # Army buildings
    "clan castle": {3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1, 12: 1, 13: 1, 14: 1, 15: 1, 16: 1, 17: 1, 18: 1},
    "army camp": {1: 1, 2: 1, 3: 2, 4: 2, 5: 2, 6: 3, 7: 4, 8: 4, 9: 4, 10: 4, 11: 4, 12: 4, 13: 4, 14: 4, 15: 4, 16: 4, 17: 4, 18: 4},
    "barracks": {1: 1, 2: 1, 3: 2, 4: 2, 5: 2, 6: 3, 7: 4, 8: 4, 9: 4, 10: 4, 11: 4, 12: 4, 13: 4, 14: 4, 15: 4, 16: 4, 17: 4, 18: 4},
    "dark barracks": {7: 1, 8: 1, 9: 2, 10: 2, 11: 2, 12: 2, 13: 2, 14: 2, 15: 2, 16: 2, 17: 2, 18: 2},
    "laboratory": {3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1, 12: 1, 13: 1, 14: 1, 15: 1, 16: 1, 17: 1, 18: 1},
    "spell factory": {5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1, 12: 1, 13: 1, 14: 1, 15: 1, 16: 1, 17: 1, 18: 1},
    "hero hall": {15: 1, 16: 1, 17: 1, 18: 1},
    "dark spell factory": {8: 1, 9: 1, 10: 1, 11: 1, 12: 1, 13: 1, 14: 1, 15: 1, 16: 1, 17: 1, 18: 1},
    "blachsmith": {15: 1, 16: 1, 17: 1, 18: 1},
    "workshop": {12: 1, 13: 1, 14: 1, 15: 1, 16: 1, 17: 1, 18: 1},
    "pet house": {14: 1, 15: 1, 16: 1, 17: 1, 18: 1},
}


# --- Building name normalization ---
NAME_NORMALIZE: Dict[str, str] = {
    "elixar collector": "Elixir Collector",
    "elixar storage": "Elixir Storage",
    "dark elixar drill": "Dark Elixir Drill",
    "dark elixar storage": "Dark Elixir Storage",
    "blachsmith": "Blacksmith",
    "x-bow": "X-Bow",
    # title-case fallback applied for others
}


def _display_name(raw: str) -> str:
    n = raw.strip().lower()
    if n in NAME_NORMALIZE:
        return NAME_NORMALIZE[n]
    return raw.strip().title()


def _is_building_name_row(row: Tuple) -> bool:
    """True if row[0] is a non-empty string and all other cells in row are empty."""
    if not row or not row[0] or not isinstance(row[0], str):
        return False
    if str(row[0]).lower().startswith("module"):
        return False  # sub-rows
    return all(c is None for c in row[1:])


def _is_header_row(row: Tuple) -> bool:
    if not row or not row[0]:
        return False
    n = _norm(row[0])
    if n not in ("level", "th level"):
        return False
    # Must contain at least one of the cost/time/th headers
    norm = [_norm(c) for c in row if c is not None]
    return any(h in norm for h in COST_HEADERS + TIME_HEADERS + TH_HEADERS)


def parse_xlsx(xlsx_path: Path) -> pd.DataFrame:
    """Parse the entire workbook into a tidy DataFrame.

    Columns:
      building (str, display name)
      raw_name (str, lowercase from xlsx)
      level (int)
      cost (int or None)
      duration_sec (int or None)
      th_required (int or None)
      resource (str)
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    records = []
    current_building: Optional[str] = None
    current_raw: Optional[str] = None
    cost_col: Optional[int] = None
    time_col: Optional[int] = None
    th_col: Optional[int] = None
    level_col: Optional[int] = None
    cap_col: Optional[int] = None
    in_data = False

    for row in rows:
        if _is_building_name_row(row):
            current_raw = str(row[0]).strip().lower()
            current_building = _display_name(row[0])
            cost_col = time_col = th_col = level_col = cap_col = None
            in_data = False
            continue

        if current_building is None:
            continue

        if _is_header_row(row):
            level_col = _find_col(row, LEVEL_HEADERS)
            cost_col = _find_col(row, COST_HEADERS)
            time_col = _find_col(row, TIME_HEADERS)
            th_col = _find_col(row, TH_HEADERS)
            cap_col = _find_col(row, CAPACITY_HEADERS)
            in_data = True
            continue

        if not in_data:
            continue

        if level_col is None or row[level_col] is None or not isinstance(row[level_col], (int, float)):
            continue

        level = int(row[level_col])
        cost = parse_cost(row[cost_col]) if cost_col is not None else None
        duration = parse_time_string(row[time_col]) if time_col is not None else None
        th_req = None
        if th_col is not None and row[th_col] is not None and isinstance(row[th_col], (int, float)):
            th_req = int(row[th_col])
        capacity = parse_cost(row[cap_col]) if cap_col is not None else None
        resource = RESOURCE_BY_BUILDING.get(current_raw, "gold")

        records.append({
            "building": current_building,
            "raw_name": current_raw,
            "level": level,
            "cost": cost,
            "duration_sec": duration,
            "th_required": th_req,
            "resource": resource,
            "capacity": capacity,
        })

    return pd.DataFrame(records)


def storage_capacity(df: pd.DataFrame, raw_name: str, level: int) -> Optional[int]:
    """Look up storage capacity for a building-level row."""
    sub = df[(df["raw_name"] == raw_name) & (df["level"] == level)]
    if sub.empty:
        return None
    val = sub.iloc[0].get("capacity")
    if pd.isna(val) or val is None:
        return None
    return int(val)


def max_level_at_th(df: pd.DataFrame, building: str, th: int) -> Optional[int]:
    """Highest level of `building` available at TH level `th`."""
    sub = df[(df["raw_name"] == building.lower()) & (df["th_required"] <= th)]
    if sub.empty:
        return None
    return int(sub["level"].max())


def instance_count(building: str, th: int) -> int:
    counts = INSTANCE_COUNT.get(building.lower())
    if not counts:
        return 0
    # find the largest known TH <= th
    avail = [k for k in counts if k <= th]
    return counts[max(avail)] if avail else 0
